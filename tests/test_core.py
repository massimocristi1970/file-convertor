import io
import unittest
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from data_io import (
    SUPPORTED_INPUT_TYPES,
    SUPPORTED_OUTPUT_TYPES,
    get_mime_type,
    read_json_bytes_safely,
    read_uploaded_file_as_df,
    read_xml_bytes_safely,
    to_export_bytes,
    to_xml_bytes,
)
from merge_utils import build_composite_key, combine_dataframes, merge_dataframes


class CoreTests(unittest.TestCase):
    def _file_obj(self, payload: bytes) -> SimpleNamespace:
        return SimpleNamespace(getvalue=lambda: payload)

    def test_xml_round_trip_preserves_columns(self) -> None:
        df = pd.DataFrame([{"Account": "0012", "Name": "Ada"}, {"Account": "0099", "Name": "Grace"}])
        xml_bytes = to_xml_bytes(df, date_format="%Y-%m-%d")
        restored = read_xml_bytes_safely(xml_bytes)
        self.assertEqual(list(restored.columns), ["Account", "Name"])
        self.assertEqual(restored.iloc[0]["Account"], "0012")

    def test_composite_key_normalises_excel_style_numbers(self) -> None:
        df = pd.DataFrame([{"AppID": "123.0", "Postcode": "AB1 2CD"}])
        keys = build_composite_key(df, ["AppID", "Postcode"])
        self.assertEqual(keys.iloc[0], "123||AB1 2CD")

    def test_merge_exclude_unmatched_keeps_only_matches(self) -> None:
        left = pd.DataFrame([{"AppID": "1", "Name": "Ada"}, {"AppID": "2", "Name": "Grace"}])
        right = pd.DataFrame([{"AppID": "1", "Score": "10"}, {"AppID": "3", "Score": "11"}])
        result = merge_dataframes(
            file_entries=[
                {"role": "Applications", "df": left, "key_cols": ["AppID"], "duplicate_strategy": "Keep first"},
                {"role": "Scores", "df": right, "key_cols": ["AppID"], "duplicate_strategy": "Keep first"},
            ],
            base_role="Applications",
            join_type="left",
            exclude_unmatched=True,
            delimiter=",",
            encoding="utf-8",
            quoting=0,
            escapechar_enabled=False,
            date_format="%Y-%m-%d",
        )
        merged = result["merged"]
        self.assertEqual(len(merged), 1)
        self.assertEqual(merged.iloc[0]["AppID"], "1")

    def test_combine_dataframes_appends_rows_and_adds_source_file(self) -> None:
        first = pd.DataFrame([{"Account": "0012", "Name": "Ada"}])
        second = pd.DataFrame([{"Account": "0099", "Name": "Grace"}])
        result = combine_dataframes(
            file_entries=[
                {"name": "data.xlsx", "df": first},
                {"name": "data (1).xlsx", "df": second},
            ],
            schema_mode="Strict same columns",
            add_source_file=True,
            source_column_name="SourceFile",
        )
        combined = result["combined"]
        self.assertEqual(len(combined), 2)
        self.assertEqual(list(combined.columns), ["Account", "Name", "SourceFile"])
        self.assertEqual(combined.iloc[0]["SourceFile"], "data.xlsx")
        self.assertEqual(combined.iloc[1]["SourceFile"], "data (1).xlsx")

    def test_combine_dataframes_rejects_schema_mismatch_in_strict_mode(self) -> None:
        first = pd.DataFrame([{"Account": "0012", "Name": "Ada"}])
        second = pd.DataFrame([{"Account": "0099", "FullName": "Grace"}])
        with self.assertRaisesRegex(RuntimeError, "does not match the first file schema"):
            combine_dataframes(
                file_entries=[
                    {"name": "data.xlsx", "df": first},
                    {"name": "data (1).xlsx", "df": second},
                ],
                schema_mode="Strict same columns",
                add_source_file=False,
                source_column_name="SourceFile",
            )

    def test_combine_dataframes_supports_union_columns(self) -> None:
        first = pd.DataFrame([{"Account": "0012", "Name": "Ada"}])
        second = pd.DataFrame([{"Account": "0099", "Score": "11"}])
        result = combine_dataframes(
            file_entries=[
                {"name": "data.xlsx", "df": first},
                {"name": "data (1).xlsx", "df": second},
            ],
            schema_mode="Union columns",
            add_source_file=False,
            source_column_name="SourceFile",
        )
        combined = result["combined"].fillna("")
        self.assertEqual(list(combined.columns), ["Account", "Name", "Score"])
        self.assertEqual(combined.to_dict(orient="records"), [
            {"Account": "0012", "Name": "Ada", "Score": ""},
            {"Account": "0099", "Name": "", "Score": "11"},
        ])

    def test_txt_plain_text_reads_as_value_column(self) -> None:
        restored = read_uploaded_file_as_df(
            file_obj=self._file_obj(b"Ada\nGrace\n"),
            file_type="txt",
            sheet_name=None,
            header_row=1,
            formula_mode="Cached values (recommended)",
            drop_empty=True,
            text_parse_mode="Delimited",
            text_delimiter=None,
        )
        self.assertEqual(restored.to_dict(orient="records"), [{"value": "Ada"}, {"value": "Grace"}])

    def test_json_round_trip_supports_all_export_orientations(self) -> None:
        df = pd.DataFrame([{"Account": "0012", "Name": "Ada"}, {"Account": "0099", "Name": "Grace"}])

        for orient in ("records", "split", "index", "columns"):
            with self.subTest(orient=orient):
                json_bytes = to_export_bytes(
                    df=df,
                    output_type="json",
                    delimiter=",",
                    encoding="utf-8",
                    quoting=0,
                    escapechar_enabled=False,
                    date_format="%Y-%m-%d",
                    json_orient=orient,
                )
                restored = read_json_bytes_safely(json_bytes).fillna("").astype(str)
                self.assertEqual(restored.to_dict(orient="records"), df.astype(str).to_dict(orient="records"))

    def test_every_supported_output_type_exports_bytes(self) -> None:
        df = pd.DataFrame([{"Account": "0012", "Name": "Ada"}])

        for output_type in SUPPORTED_OUTPUT_TYPES:
            with self.subTest(output_type=output_type):
                out = to_export_bytes(
                    df=df,
                    output_type=output_type,
                    delimiter="\t" if output_type == "tsv" else ",",
                    encoding="utf-8",
                    quoting=0,
                    escapechar_enabled=False,
                    date_format="%Y-%m-%d",
                    json_orient="records",
                )
                self.assertIsInstance(out, bytes)
                self.assertGreater(len(out), 0)
                self.assertTrue(get_mime_type(output_type))

    def test_xlsx_export_keeps_datetime_columns_as_excel_dates(self) -> None:
        df = pd.DataFrame([{"EventDate": pd.Timestamp("2026-04-21"), "Name": "Ada"}])
        out = to_export_bytes(
            df=df,
            output_type="xlsx",
            delimiter=",",
            encoding="utf-8",
            quoting=0,
            escapechar_enabled=False,
            date_format="%Y-%m-%d",
        )
        workbook = load_workbook(io.BytesIO(out))
        cell = workbook["Sheet1"]["A2"]
        self.assertIsNotNone(cell.value)
        self.assertNotIsInstance(cell.value, str)
        self.assertEqual(cell.number_format, "yyyy-mm-dd")

    def test_xlsx_export_formats_parseable_date_columns_as_excel_dates(self) -> None:
        df = pd.DataFrame([{"DateOfBirth": "21/04/2026", "Name": "Ada"}])
        out = to_export_bytes(
            df=df,
            output_type="xlsx",
            delimiter=",",
            encoding="utf-8",
            quoting=0,
            escapechar_enabled=False,
            date_format="%Y-%m-%d",
        )
        workbook = load_workbook(io.BytesIO(out))
        cell = workbook["Sheet1"]["A2"]
        self.assertIsNotNone(cell.value)
        self.assertNotIsInstance(cell.value, str)
        self.assertEqual(cell.number_format, "yyyy-mm-dd")

    def test_supported_input_types_are_covered_by_readers(self) -> None:
        fixtures = {
            "csv": b"Account,Name\n0012,Ada\n",
            "tsv": b"Account\tName\n0012\tAda\n",
            "txt": b"Account,Name\n0012,Ada\n",
            "json": b'[{"Account":"0012","Name":"Ada"}]',
            "xml": b'<?xml version="1.0" encoding="utf-8"?><rows><row><field name="Account">0012</field><field name="Name">Ada</field></row></rows>',
        }

        for input_type in SUPPORTED_INPUT_TYPES:
            with self.subTest(input_type=input_type):
                if input_type == "xlsx":
                    payload = to_export_bytes(
                        df=pd.DataFrame([{"Account": "0012", "Name": "Ada"}]),
                        output_type="xlsx",
                        delimiter=",",
                        encoding="utf-8",
                        quoting=0,
                        escapechar_enabled=False,
                        date_format="%Y-%m-%d",
                    )
                elif input_type == "parquet":
                    payload = to_export_bytes(
                        df=pd.DataFrame([{"Account": "0012", "Name": "Ada"}]),
                        output_type="parquet",
                        delimiter=",",
                        encoding="utf-8",
                        quoting=0,
                        escapechar_enabled=False,
                        date_format="%Y-%m-%d",
                    )
                else:
                    payload = fixtures[input_type]

                restored = read_uploaded_file_as_df(
                    file_obj=self._file_obj(payload),
                    file_type=input_type,
                    sheet_name="Sheet1" if input_type == "xlsx" else None,
                    header_row=1,
                    formula_mode="Cached values (recommended)",
                    drop_empty=True,
                    text_parse_mode="Delimited",
                    text_delimiter=None,
                )
                self.assertFalse(restored.empty)
                self.assertIn("Account", restored.columns)


if __name__ == "__main__":
    unittest.main()
