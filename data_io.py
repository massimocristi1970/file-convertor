import csv
import io
import json
import re
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional

import pandas as pd

try:
    import openpyxl
except ImportError:
    openpyxl = None


SUPPORTED_INPUT_TYPES = ["xlsx", "csv", "tsv", "txt", "json", "xml", "parquet"]
SUPPORTED_OUTPUT_TYPES = ["csv", "tsv", "txt", "xlsx", "json", "xml", "parquet"]
TEXT_LIKE_TYPES = {"csv", "tsv", "txt"}
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def safe_filename(name: str, max_len: int = 80) -> str:
    name = name.strip()
    name = re.sub(r"[^\w\-. ]+", "_", name)
    name = re.sub(r"\s+", " ", name)
    name = name.strip(" ._")
    if not name:
        name = "sheet"
    return name[:max_len]


def parse_csv_columns(raw: str) -> List[str]:
    if not raw:
        return []
    return [part.strip() for part in raw.split(",") if part.strip()]


def delimiter_from_choice(choice: str, custom: str) -> str:
    mapping = {
        "Comma (,)": ",",
        "Semicolon (;)": ";",
        "Tab (\\t)": "\t",
        "Pipe (|)": "|",
        "Space": " ",
        "Custom": custom if custom else ",",
    }
    return mapping.get(choice, ",")


def quoting_from_choice(choice: str) -> int:
    mapping = {
        "Minimal (default)": csv.QUOTE_MINIMAL,
        "All fields": csv.QUOTE_ALL,
        "Non-numeric": csv.QUOTE_NONNUMERIC,
        "None": csv.QUOTE_NONE,
    }
    return mapping.get(choice, csv.QUOTE_MINIMAL)


def detect_file_type(name: str) -> str:
    lowered = (name or "").lower()
    if "." not in lowered:
        raise RuntimeError("Uploaded file has no extension.")
    ext = lowered.rsplit(".", 1)[1]
    if ext not in SUPPORTED_INPUT_TYPES:
        raise RuntimeError(f"Unsupported file type: .{ext}")
    return ext


def read_text_bytes(file_bytes: bytes) -> str:
    last_err: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except Exception as exc:
            last_err = exc
    raise RuntimeError(f"Could not decode text input. Last error: {last_err}")


def load_workbook_bytes(xlsx_bytes: bytes, data_only: bool) -> "openpyxl.Workbook":
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Install it with: pip install openpyxl")
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=data_only, read_only=False)


def list_xlsx_sheets(xlsx_bytes: bytes, data_only: bool) -> List[str]:
    wb = load_workbook_bytes(xlsx_bytes, data_only=data_only)
    return list(wb.sheetnames)


def read_sheet_as_dataframe(
    xlsx_bytes: bytes,
    sheet_name: str,
    header_row: int,
    formula_mode: str,
    drop_empty: bool,
) -> pd.DataFrame:
    df = pd.read_excel(
        io.BytesIO(xlsx_bytes),
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )

    if drop_empty:
        df = df.dropna(how="all").dropna(axis=1, how="all")

    if formula_mode == "Formula strings":
        wb = load_workbook_bytes(xlsx_bytes, data_only=False)
        ws = wb[sheet_name]
        values = list(ws.values)
        if not values:
            return pd.DataFrame()

        hdr_idx = header_row - 1
        if hdr_idx >= len(values):
            return pd.DataFrame()

        headers = list(values[hdr_idx])
        data_rows = values[hdr_idx + 1 :]
        clean_headers: List[str] = []
        seen: Dict[str, int] = {}
        for header in headers:
            header_name = str(header).strip() if header is not None else ""
            if not header_name:
                header_name = "Unnamed"
            seen[header_name] = seen.get(header_name, -1) + 1
            clean_headers.append(header_name if seen[header_name] == 0 else f"{header_name}.{seen[header_name]}")

        df2 = pd.DataFrame(data_rows, columns=clean_headers)
        if drop_empty:
            df2 = df2.dropna(how="all").dropna(axis=1, how="all")
        return df2

    return df


def force_columns_to_text(df: pd.DataFrame, col_names: List[str]) -> pd.DataFrame:
    if not col_names:
        return df
    out = df.copy()
    for col in col_names:
        if col in out.columns:
            out[col] = out[col].map(lambda x: "" if pd.isna(x) else str(x))
    return out


def normalise_merge_key(series: pd.Series) -> pd.Series:
    out = series.map(lambda v: "" if pd.isna(v) else str(v).strip())
    return out.str.replace(r"\.0$", "", regex=True)


def normalise_dates(df: pd.DataFrame, date_format: str) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime(date_format)
    return out


def _python_date_format_to_excel(date_format: str) -> str:
    mapping = {
        "%Y": "yyyy",
        "%y": "yy",
        "%m": "mm",
        "%d": "dd",
    }
    excel_format = str(date_format or "%Y-%m-%d")
    for token, replacement in mapping.items():
        excel_format = excel_format.replace(token, replacement)
    return excel_format


def prepare_xlsx_dates(df: pd.DataFrame) -> tuple[pd.DataFrame, List[str]]:
    out = df.copy()
    date_columns: List[str] = []

    for col in out.columns:
        series = out[col]
        column_name = str(col).strip().lower()

        if pd.api.types.is_datetime64_any_dtype(series):
            out[col] = pd.to_datetime(series, errors="coerce").dt.tz_localize(None)
            date_columns.append(col)
            continue

        if not (pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series)):
            continue
        if not any(token in column_name for token in ("date", "dob", "birth", "expiry", "issued", "created", "updated")):
            continue

        non_blank = series[~pd.isna(series)].map(lambda value: str(value).strip())
        non_blank = non_blank[non_blank != ""]
        if non_blank.empty:
            continue

        parsed = pd.to_datetime(non_blank, errors="coerce", dayfirst=True)
        if parsed.notna().all():
            out[col] = pd.to_datetime(series, errors="coerce", dayfirst=True)
            date_columns.append(col)

    return out, date_columns


def sniff_delimiter(sample: str, fallback: str = ",") -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        return dialect.delimiter
    except Exception:
        return fallback


def read_delimited_bytes_safely(file_bytes: bytes, delimiter: Optional[str], header_row: int) -> pd.DataFrame:
    last_err: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            if delimiter is None:
                sample = file_bytes[:4096].decode(enc)
                sep = sniff_delimiter(sample)
            else:
                sep = delimiter
            return pd.read_csv(io.BytesIO(file_bytes), encoding=enc, sep=sep, header=header_row - 1)
        except Exception as exc:
            last_err = exc
    raise RuntimeError(f"Could not read delimited text with common encodings. Last error: {last_err}")


def read_plain_text_lines(file_bytes: bytes) -> pd.DataFrame:
    text = read_text_bytes(file_bytes)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return pd.DataFrame({"value": lines})


def should_treat_txt_as_plain_text(file_bytes: bytes, delimiter: Optional[str], header_row: int) -> bool:
    if delimiter is not None or header_row != 1:
        return False

    text = read_text_bytes(file_bytes)
    non_empty_lines = [line for line in text.splitlines() if line.strip()]
    if len(non_empty_lines) <= 1:
        return True

    sample = "\n".join(non_empty_lines[:5])
    detected = sniff_delimiter(sample, fallback="")
    if not detected:
        return True

    split_counts = [len(line.split(detected)) for line in non_empty_lines[:5]]
    return max(split_counts, default=1) <= 1


def read_fixed_width_bytes_safely(file_bytes: bytes, header_row: int) -> pd.DataFrame:
    text = read_text_bytes(file_bytes)
    return pd.read_fwf(io.StringIO(text), header=header_row - 1)


def _records_from_json(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return [item if isinstance(item, dict) else {"value": item} for item in data]
    if isinstance(data, dict):
        list_values = [value for value in data.values() if isinstance(value, list)]
        if len(list_values) == 1:
            return [item if isinstance(item, dict) else {"value": item} for item in list_values[0]]
        return [data]
    return [{"value": data}]


def read_json_bytes_safely(file_bytes: bytes) -> pd.DataFrame:
    text = read_text_bytes(file_bytes)
    data = json.loads(text)
    if isinstance(data, dict) and {"columns", "data"}.issubset(data.keys()):
        return pd.DataFrame(data["data"], columns=data.get("columns"), index=data.get("index"))
    if isinstance(data, dict) and data and all(isinstance(value, dict) for value in data.values()):
        numeric_outer_keys = all(str(key).isdigit() for key in data.keys())
        numeric_inner_keys = all(all(str(inner_key).isdigit() for inner_key in value.keys()) for value in data.values())
        if numeric_outer_keys:
            return pd.DataFrame.from_dict(data, orient="index")
        if numeric_inner_keys:
            return pd.DataFrame.from_dict(data, orient="columns")
    if isinstance(data, dict) and data and all(isinstance(value, list) for value in data.values()):
        return pd.DataFrame(data)
    return pd.json_normalize(_records_from_json(data))


def read_xml_bytes_safely(file_bytes: bytes) -> pd.DataFrame:
    root = ET.fromstring(file_bytes)
    rows: List[Dict[str, Any]] = []
    for child in list(root):
        record: Dict[str, Any] = {}
        field_children = [node for node in list(child) if node.tag == "field" and "name" in node.attrib]
        if field_children:
            for field in field_children:
                record[field.attrib["name"]] = field.text or ""
        else:
            for field in list(child):
                record[field.tag] = field.text or ""
        if record:
            rows.append(record)
    return pd.DataFrame(rows)


def looks_like_xml_text(file_bytes: bytes) -> bool:
    try:
        text = read_text_bytes(file_bytes).lstrip()
    except Exception:
        return False
    return text.startswith("<?xml") or text.startswith("<")


def read_parquet_bytes_safely(file_bytes: bytes) -> pd.DataFrame:
    return pd.read_parquet(io.BytesIO(file_bytes))


def read_uploaded_file_as_df(
    file_obj: Any,
    file_type: str,
    sheet_name: Optional[str],
    header_row: int,
    formula_mode: str,
    drop_empty: bool,
    text_parse_mode: str = "Delimited",
    text_delimiter: Optional[str] = None,
) -> pd.DataFrame:
    file_bytes = file_obj.getvalue()

    if file_type == "xlsx":
        selected_sheet = sheet_name
        if selected_sheet is None:
            sheets = list_xlsx_sheets(file_bytes, data_only=(formula_mode == "Cached values (recommended)"))
            selected_sheet = sheets[0] if sheets else None
            if selected_sheet is None:
                return pd.DataFrame()
        df = read_sheet_as_dataframe(file_bytes, selected_sheet, header_row, formula_mode, drop_empty)
    elif file_type in {"csv", "tsv", "txt"}:
        parse_mode = text_parse_mode or "Delimited"
        delimiter = text_delimiter
        if file_type == "tsv" and delimiter is None:
            delimiter = "\t"
        if file_type == "csv" and delimiter is None:
            delimiter = ","
        if file_type == "txt" and looks_like_xml_text(file_bytes):
            try:
                df = read_xml_bytes_safely(file_bytes)
            except Exception:
                if parse_mode == "Fixed width":
                    df = read_fixed_width_bytes_safely(file_bytes, header_row)
                else:
                    df = read_delimited_bytes_safely(file_bytes, delimiter, header_row)
        elif file_type == "txt" and parse_mode != "Fixed width" and should_treat_txt_as_plain_text(file_bytes, delimiter, header_row):
            df = read_plain_text_lines(file_bytes)
        elif parse_mode == "Fixed width":
            df = read_fixed_width_bytes_safely(file_bytes, header_row)
        else:
            df = read_delimited_bytes_safely(file_bytes, delimiter, header_row)
    elif file_type == "json":
        df = read_json_bytes_safely(file_bytes)
    elif file_type == "xml":
        df = read_xml_bytes_safely(file_bytes)
    elif file_type == "parquet":
        df = read_parquet_bytes_safely(file_bytes)
    else:
        raise RuntimeError(f"Unsupported file type: {file_type}")

    if drop_empty:
        df = df.dropna(how="all").dropna(axis=1, how="all")
    return df


def to_csv_bytes(
    df: pd.DataFrame,
    delimiter: str,
    encoding: str,
    quoting: int,
    escapechar_enabled: bool,
    date_format: str,
) -> bytes:
    out = normalise_dates(df, date_format=date_format)
    buf = io.StringIO()
    out.to_csv(
        buf,
        index=False,
        sep=delimiter,
        encoding=None,
        quoting=quoting,
        escapechar="\\" if escapechar_enabled else None,
        quotechar='"',
        lineterminator="\n",
    )
    return buf.getvalue().encode(encoding, errors="replace")


def to_xlsx_bytes(df: pd.DataFrame, date_format: str) -> bytes:
    buf = io.BytesIO()
    excel_df, date_columns = prepare_xlsx_dates(df)
    excel_date_format = _python_date_format_to_excel(date_format)
    with pd.ExcelWriter(
        buf,
        engine="openpyxl",
        date_format=excel_date_format,
        datetime_format=excel_date_format,
    ) as writer:
        excel_df.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.book["Sheet1"]
        for col_idx, column_name in enumerate(excel_df.columns, start=1):
            if column_name not in date_columns:
                continue
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_idx, column=col_idx).number_format = excel_date_format
    buf.seek(0)
    return buf.getvalue()


def to_json_bytes(df: pd.DataFrame, orient: str, date_format: str) -> bytes:
    out = normalise_dates(df, date_format=date_format)
    return out.to_json(orient=orient, date_format="iso", force_ascii=False).encode("utf-8", errors="replace")


def to_xml_bytes(df: pd.DataFrame, date_format: str, root_name: str = "rows", row_name: str = "row") -> bytes:
    out = normalise_dates(df, date_format=date_format).copy()
    root = ET.Element(root_name)
    for record in out.to_dict(orient="records"):
        row_el = ET.SubElement(root, row_name)
        for column, value in record.items():
            field_el = ET.SubElement(row_el, "field", name=str(column))
            field_el.text = "" if pd.isna(value) else str(value)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def to_parquet_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_parquet(buf, index=False)
    buf.seek(0)
    return buf.getvalue()


def to_export_bytes(
    df: pd.DataFrame,
    output_type: str,
    delimiter: str,
    encoding: str,
    quoting: int,
    escapechar_enabled: bool,
    date_format: str,
    json_orient: str = "records",
    xml_root: str = "rows",
    xml_row: str = "row",
) -> bytes:
    if output_type in TEXT_LIKE_TYPES:
        return to_csv_bytes(df, delimiter, encoding, quoting, escapechar_enabled, date_format)
    if output_type == "xlsx":
        return to_xlsx_bytes(df, date_format)
    if output_type == "json":
        return to_json_bytes(df, json_orient, date_format)
    if output_type == "xml":
        return to_xml_bytes(df, date_format, root_name=xml_root, row_name=xml_row)
    if output_type == "parquet":
        return to_parquet_bytes(df)
    raise RuntimeError(f"Unsupported output type: {output_type}")


def get_mime_type(output_type: str) -> str:
    mapping = {
        "csv": "text/csv",
        "tsv": "text/tab-separated-values",
        "txt": "text/plain",
        "xlsx": EXCEL_MIME,
        "json": "application/json",
        "xml": "application/xml",
        "parquet": "application/octet-stream",
    }
    return mapping[output_type]
