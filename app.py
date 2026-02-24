import io
import json
import re
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
import streamlit as st

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Short, direct name (not too long)
APP_TITLE = "Data Mapper"


# ---------- Helpers ----------
def safe_filename(name: str, max_len: int = 80) -> str:
    name = name.strip()
    name = re.sub(r"[^\w\-. ]+", "_", name)
    name = re.sub(r"\s+", " ", name)
    name = name.strip(" ._")
    if not name:
        name = "sheet"
    return name[:max_len]


def parse_force_text_columns(raw: str) -> List[str]:
    """
    User can enter: "colA, colB, AccountNumber"
    We'll treat them as column names after header row is read.
    """
    if not raw:
        return []
    parts = [p.strip() for p in raw.split(",")]
    return [p for p in parts if p]


def delimiter_from_choice(choice: str, custom: str) -> str:
    mapping = {
        "Comma (,)": ",",
        "Semicolon (;)": ";",
        "Tab (\\t)": "\t",
        "Pipe (|)": "|",
        "Custom": custom if custom else ",",
    }
    return mapping.get(choice, ",")


def quoting_from_choice(choice: str) -> int:
    import csv

    mapping = {
        "Minimal (default)": csv.QUOTE_MINIMAL,
        "All fields": csv.QUOTE_ALL,
        "Non-numeric": csv.QUOTE_NONNUMERIC,
        "None": csv.QUOTE_NONE,
    }
    return mapping.get(choice, csv.QUOTE_MINIMAL)


def load_workbook_bytes(xlsx_bytes: bytes, data_only: bool) -> "openpyxl.Workbook":
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Install it with: pip install openpyxl")
    bio = io.BytesIO(xlsx_bytes)
    # read_only=False to allow pandas to read properly; data_only controls formulas vs cached values
    return openpyxl.load_workbook(bio, data_only=data_only, read_only=False)


def read_sheet_as_dataframe(
    xlsx_bytes: bytes,
    sheet_name: str,
    header_row: int,
    formula_mode: str,
    drop_empty: bool,
) -> pd.DataFrame:
    """
    formula_mode:
      - "Cached values (recommended)" -> data_only=True
      - "Formula strings"             -> data_only=False
    """
    df = pd.read_excel(
        io.BytesIO(xlsx_bytes),
        sheet_name=sheet_name,
        header=header_row - 1,  # pandas is 0-indexed
        engine="openpyxl",
    )

    if drop_empty:
        df = df.dropna(how="all").dropna(axis=1, how="all")

    # To truly export formula strings we need openpyxl cell inspection.
    if formula_mode == "Formula strings":
        wb = load_workbook_bytes(xlsx_bytes, data_only=False)
        ws = wb[sheet_name]

        values = list(ws.values)
        if not values:
            return pd.DataFrame()

        hdr_idx = header_row - 1
        if hdr_idx >= len(values):
            return pd.DataFrame()

        headers = list(values[hdr_idx])
        data_rows = values[hdr_idx + 1 :]

        # Make unique header names
        clean_headers: List[str] = []
        seen: Dict[str, int] = {}
        for h in headers:
            h_str = str(h).strip() if h is not None else ""
            if not h_str:
                h_str = "Unnamed"
            if h_str in seen:
                seen[h_str] += 1
                h2 = f"{h_str}.{seen[h_str]}"
            else:
                seen[h_str] = 0
                h2 = h_str
            clean_headers.append(h2)

        df2 = pd.DataFrame(data_rows, columns=clean_headers)
        if drop_empty:
            df2 = df2.dropna(how="all").dropna(axis=1, how="all")
        return df2

    return df


def force_columns_to_text(df: pd.DataFrame, col_names: List[str]) -> pd.DataFrame:
    """
    Forces specific columns (if present) to string and preserves leading zeros.
    """
    if not col_names:
        return df
    df = df.copy()
    for col in col_names:
        if col in df.columns:
            df[col] = df[col].map(lambda x: "" if pd.isna(x) else str(x))
    return df


def normalise_merge_key(series: pd.Series) -> pd.Series:
    """
    Normalise merge keys to safe strings.
    - trims whitespace
    - removes trailing ".0" (common when IDs were numeric in Excel)
    - preserves leading zeros (because we keep as strings)
    """
    s = series.map(lambda v: "" if pd.isna(v) else str(v).strip())
    s = s.str.replace(r"\.0$", "", regex=True)
    return s


def normalise_dates(df: pd.DataFrame, date_format: str) -> pd.DataFrame:
    """
    Format datetime-like columns to consistent string format for CSV output.
    """
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime(date_format)
    return df


def to_csv_bytes(
    df: pd.DataFrame,
    delimiter: str,
    encoding: str,
    quoting: int,
    escapechar_enabled: bool,
    date_format: str,
) -> bytes:
    import csv  # noqa: F401

    df_out = normalise_dates(df, date_format=date_format)

    buf = io.StringIO()
    df_out.to_csv(
        buf,
        index=False,
        sep=delimiter,
        encoding=None,  # we encode after
        quoting=quoting,
        escapechar="\\" if escapechar_enabled else None,
        quotechar='"',
        lineterminator="\n",
    )
    return buf.getvalue().encode(encoding, errors="replace")


def read_csv_bytes_safely(file_bytes: bytes) -> pd.DataFrame:
    """
    Best-effort CSV read with encoding fallbacks.
    """
    last_err: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return pd.read_csv(io.BytesIO(file_bytes), encoding=enc)
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Could not read CSV with common encodings. Last error: {last_err}")


def read_uploaded_file_as_df(
    file_obj: Any,
    sheet_name: Optional[str],
    header_row: int,
    formula_mode: str,
    drop_empty: bool,
) -> pd.DataFrame:
    """
    Supports .xlsx and .csv
    """
    name = (getattr(file_obj, "name", "") or "").lower()

    if name.endswith(".csv"):
        df = read_csv_bytes_safely(file_obj.getvalue())
        if drop_empty:
            df = df.dropna(how="all").dropna(axis=1, how="all")
        return df

    if not name.endswith(".xlsx"):
        raise RuntimeError("Unsupported file type. Upload .xlsx or .csv")

    xlsx_bytes = file_obj.getvalue()
    if sheet_name is None:
        wb_tmp = load_workbook_bytes(xlsx_bytes, data_only=(formula_mode == "Cached values (recommended)"))
        sheet_name = wb_tmp.sheetnames[0] if wb_tmp.sheetnames else None
        if sheet_name is None:
            return pd.DataFrame()

    return read_sheet_as_dataframe(
        xlsx_bytes=xlsx_bytes,
        sheet_name=sheet_name,
        header_row=header_row,
        formula_mode=formula_mode,
        drop_empty=drop_empty,
    )


# ---------- Transform engine (future-proof, parameterised) ----------
POSTCODE_RE = re.compile(r"\b([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})\b", flags=re.IGNORECASE)


def _as_str(x: Any) -> str:
    return "" if pd.isna(x) else str(x)


def tf_none(x: str, p: Dict[str, Any]) -> str:
    return x


def tf_text_force(x: str, p: Dict[str, Any]) -> str:
    # Useful to prevent Excel/scientific notation display issues; keeps value as text.
    return x


def tf_uk_postcode(x: str, p: Dict[str, Any]) -> str:
    m = POSTCODE_RE.search(x.upper())
    return m.group(1).strip() if m else ""


def tf_first_line(x: str, p: Dict[str, Any]) -> str:
    s = x.strip()
    if not s:
        return ""
    parts = [t.strip() for t in s.split(",")]
    for t in parts:
        if t:
            return t
    return s


def tf_uk_mobile_add44(x: str, p: Dict[str, Any]) -> str:
    # Remove whitespace
    s = re.sub(r"\s+", "", x.strip())

    if not s:
        return ""

    # Remove trailing .0 caused by float conversion
    if s.endswith(".0"):
        s = s[:-2]

    # Remove all non-digits
    digits = re.sub(r"\D", "", s)

    # Already starts with 44
    if digits.startswith("44"):
        return digits

    # Starts with 0 → convert to 44
    if digits.startswith("0"):
        return "44" + digits[1:]

    # Otherwise assume missing country code
    return "44" + digits


def tf_digits_last_n(x: str, p: Dict[str, Any]) -> str:
    n = int(p.get("n", 4))
    digits = re.findall(r"\d", x)
    return "".join(digits[-n:]) if len(digits) >= n else ""


def tf_extract_regex(x: str, p: Dict[str, Any]) -> str:
    pattern = str(p.get("pattern", "") or "")
    if not pattern:
        return ""
    flags = re.IGNORECASE if bool(p.get("ignore_case", True)) else 0
    m = re.search(pattern, x, flags=flags)
    if not m:
        return ""
    group = int(p.get("group", 1))
    try:
        return (m.group(group) or "").strip()
    except Exception:
        return ""


def tf_split_take(x: str, p: Dict[str, Any]) -> str:
    delim = str(p.get("delim", ","))
    idx = int(p.get("index", 0))
    parts = [t.strip() for t in x.split(delim)]
    if not parts:
        return ""
    if idx < 0:
        idx = len(parts) + idx
    return parts[idx] if 0 <= idx < len(parts) else ""


def tf_prefix_if_missing(x: str, p: Dict[str, Any]) -> str:
    prefix = str(p.get("prefix", "") or "")
    if not prefix:
        return x
    return x if x.startswith(prefix) else prefix + x


def tf_suffix(x: str, p: Dict[str, Any]) -> str:
    suffix = str(p.get("suffix", "") or "")
    return x + suffix if suffix else x


def tf_regex_replace(x: str, p: Dict[str, Any]) -> str:
    pattern = str(p.get("pattern", "") or "")
    repl = str(p.get("repl", "") or "")
    if not pattern:
        return x
    flags = re.IGNORECASE if bool(p.get("ignore_case", True)) else 0
    return re.sub(pattern, repl, x, flags=flags)


TRANSFORM_FUNCS = {
    "None": tf_none,
    "Text (force)": tf_text_force,
    # Presets (your examples)
    "UK Postcode (extract)": tf_uk_postcode,
    "Address first line (before comma)": tf_first_line,
    "UK mobile → 44": tf_uk_mobile_add44,
    "Digits: keep last N": tf_digits_last_n,
    # Generic / future-proof
    "Extract by regex": tf_extract_regex,
    "Split + take part": tf_split_take,
    "Prefix if missing": tf_prefix_if_missing,
    "Suffix": tf_suffix,
    "Regex replace": tf_regex_replace,
}


def apply_transform(series: pd.Series, tf_name: str, params: Dict[str, Any]) -> pd.Series:
    fn = TRANSFORM_FUNCS.get(tf_name, tf_none)
    return series.map(lambda v: fn(_as_str(v), params))


# ---------- Template helpers ----------
def apply_template_to_defaults(template: Optional[dict]) -> dict:
    t = template or {}
    return {
        "version": int(t.get("version", 1)),
        "join_type": t.get("join_type", "Left (recommended)"),
        "base_role": t.get("base_role", ""),
        "merge_keys_by_role": t.get("merge_keys_by_role", {}),
        "output_spec": t.get("output_spec", []),
    }


def build_template_payload(
    join_type: str,
    base_role: str,
    merge_keys_by_role: Dict[str, str],
    out_rows: List[Dict[str, Any]],
) -> dict:
    return {
        "version": 1,
        "join_type": join_type,
        "base_role": base_role,
        "merge_keys_by_role": merge_keys_by_role,
        "output_spec": out_rows,
    }


# ---------- UI ----------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

st.markdown(
    """
Upload files and either:

- **Convert** a single XLSX to CSV / ZIP (all sheets), or
- **Merge + Map + Transform** across multiple uploads and export a clean output file.
"""
)

with st.sidebar:
    st.header("Mode")
    app_mode = st.radio(
        "Choose workflow",
        options=["Simple XLSX → CSV", "Merge + Map + Transform"],
        index=0,
    )

    st.divider()
    st.header("Export options")

    header_row = st.number_input("Header row (1 = first row)", min_value=1, max_value=100, value=1, step=1)

    formula_mode = st.selectbox(
        "Formulas",
        options=["Cached values (recommended)", "Formula strings"],
        index=0,
        help="Cached values are what Excel last calculated and saved. Formula strings exports '=SUM(...)' etc.",
    )

    drop_empty = st.checkbox("Drop completely empty rows/columns", value=True)

    st.divider()
    st.subheader("CSV formatting")

    delim_choice = st.selectbox("Delimiter", ["Comma (,)", "Semicolon (;)", "Tab (\\t)", "Pipe (|)", "Custom"], index=0)
    custom_delim = st.text_input("Custom delimiter", value=",") if delim_choice == "Custom" else ""
    delimiter = delimiter_from_choice(delim_choice, custom_delim)

    encoding = st.selectbox("Encoding", ["utf-8", "utf-8-sig (Excel-friendly)", "cp1252"], index=1)
    encoding_map = {"utf-8": "utf-8", "utf-8-sig (Excel-friendly)": "utf-8-sig", "cp1252": "cp1252"}

    quoting_choice = st.selectbox("Quoting", ["Minimal (default)", "All fields", "Non-numeric", "None"], index=0)
    quoting = quoting_from_choice(quoting_choice)

    escapechar_enabled = st.checkbox(
        "Enable escape character (\\) (useful if quoting=None)",
        value=(quoting_choice == "None"),
    )

    date_format = st.text_input("Date format", value="%Y-%m-%d", help="Python strftime format")

    st.divider()
    st.subheader("Data fidelity")

    force_text_raw = st.text_area(
        "Force these columns to TEXT (preserve IDs/leading zeros)\nComma-separated column names:",
        value="",
        placeholder="e.g. AccountNumber, SortCode, Postcode, Mobile",
    )
    force_text_cols = parse_force_text_columns(force_text_raw)

    preview_rows = st.slider("Preview rows", min_value=5, max_value=200, value=25, step=5)

    st.divider()
    st.subheader("Mapping template (optional)")
    template_file = st.file_uploader("Load template (.json)", type=["json"], key="template_json")
    loaded_template = None
    tmpl_defaults = apply_template_to_defaults(None)
    if template_file:
        try:
            loaded_template = json.loads(template_file.getvalue().decode("utf-8"))
            tmpl_defaults = apply_template_to_defaults(loaded_template)
            st.success("Template loaded.")
        except Exception as e:
            st.error(f"Failed to load template: {e}")


# ---------------------------
# Mode 1: Simple XLSX → CSV
# ---------------------------
if app_mode == "Simple XLSX → CSV":
    st.markdown("This keeps your original converter behaviour intact.")

    uploaded = st.file_uploader("Upload .xlsx", type=["xlsx"])

    if not uploaded:
        st.info("Upload an .xlsx file to begin.")
        st.stop()

    export_mode = st.radio(
        "Export mode",
        options=["Single sheet → CSV", "All sheets → ZIP of CSVs"],
        index=0,
        horizontal=True,
    )

    xlsx_bytes = uploaded.getvalue()

    try:
        wb_tmp = load_workbook_bytes(xlsx_bytes, data_only=(formula_mode == "Cached values (recommended)"))
        sheet_names = wb_tmp.sheetnames
    except Exception as e:
        st.error(f"Could not read workbook. Error: {e}")
        st.stop()

    if not sheet_names:
        st.error("No worksheets found in the uploaded file.")
        st.stop()

    col1, col2 = st.columns([1, 1], gap="large")

    with col1:
        st.subheader("Workbook")
        st.write(
            {
                "filename": uploaded.name,
                "size_kb": round(len(xlsx_bytes) / 1024, 1),
                "sheets": sheet_names,
            }
        )

    with col2:
        st.subheader("Preview")

        if export_mode == "Single sheet → CSV":
            sheet = st.selectbox("Select sheet", sheet_names)
            try:
                df = read_sheet_as_dataframe(
                    xlsx_bytes=xlsx_bytes,
                    sheet_name=sheet,
                    header_row=int(header_row),
                    formula_mode=formula_mode,
                    drop_empty=drop_empty,
                )
                df = force_columns_to_text(df, force_text_cols)
                st.dataframe(df.head(preview_rows), use_container_width=True)
                st.caption(f"Rows: {len(df):,} | Columns: {df.shape[1]:,}")
            except Exception as e:
                st.error(f"Failed to read sheet '{sheet}'. Error: {e}")
                st.stop()
        else:
            sheet = sheet_names[0]
            try:
                df = read_sheet_as_dataframe(
                    xlsx_bytes=xlsx_bytes,
                    sheet_name=sheet,
                    header_row=int(header_row),
                    formula_mode=formula_mode,
                    drop_empty=drop_empty,
                )
                df = force_columns_to_text(df, force_text_cols)
                st.dataframe(df.head(preview_rows), use_container_width=True)
                st.caption(f"Previewing first sheet: {sheet} | Rows: {len(df):,} | Columns: {df.shape[1]:,}")
            except Exception as e:
                st.error(f"Failed to read sheet '{sheet}'. Error: {e}")
                st.stop()

    st.divider()
    st.subheader("Download")

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    base_name = safe_filename(uploaded.name.rsplit(".", 1)[0])

    if export_mode == "Single sheet → CSV":
        out_name = f"{base_name}_{safe_filename(sheet)}_{timestamp}.csv"
        try:
            csv_bytes = to_csv_bytes(
                df=df,
                delimiter=delimiter,
                encoding=encoding_map[encoding],
                quoting=quoting,
                escapechar_enabled=escapechar_enabled,
                date_format=date_format,
            )
            st.download_button(
                label=f"Download CSV ({sheet})",
                data=csv_bytes,
                file_name=out_name,
                mime="text/csv",
            )
        except Exception as e:
            st.error(f"Failed to create CSV. Error: {e}")
    else:
        zip_name = f"{base_name}_{timestamp}_csvs.zip"
        try:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
                for s in sheet_names:
                    df_s = read_sheet_as_dataframe(
                        xlsx_bytes=xlsx_bytes,
                        sheet_name=s,
                        header_row=int(header_row),
                        formula_mode=formula_mode,
                        drop_empty=drop_empty,
                    )
                    df_s = force_columns_to_text(df_s, force_text_cols)
                    csv_s = to_csv_bytes(
                        df=df_s,
                        delimiter=delimiter,
                        encoding=encoding_map[encoding],
                        quoting=quoting,
                        escapechar_enabled=escapechar_enabled,
                        date_format=date_format,
                    )
                    csv_filename = f"{safe_filename(s)}.csv"
                    z.writestr(csv_filename, csv_s)

            zip_buf.seek(0)
            st.download_button(
                label="Download ZIP (all sheets as CSV)",
                data=zip_buf.getvalue(),
                file_name=zip_name,
                mime="application/zip",
            )
        except Exception as e:
            st.error(f"Failed to create ZIP. Error: {e}")

    st.stop()


# -----------------------------------
# Mode 2: Merge + Map + Transform
# -----------------------------------
st.markdown(
    """
### 1) Upload multiple files
Upload **.xlsx** and/or **.csv** files. You can set a **role** and **merge key** for each file.
"""
)

uploaded_files = st.file_uploader(
    "Upload files (.xlsx or .csv)",
    type=["xlsx", "csv"],
    accept_multiple_files=True,
)

if not uploaded_files:
    st.info("Upload one or more files to begin.")
    st.stop()

dfs: List[pd.DataFrame] = []
file_configs: List[Dict[str, Any]] = []

for i, f in enumerate(uploaded_files):
    st.markdown(f"#### File {i+1}: `{f.name}`")

    default_role = f"File{i+1}"
    role = st.text_input(
        f"Role name ({f.name})",
        value=default_role,
        key=f"role_{i}",
        help="Give each upload a stable label (e.g. Applications, Bureau, Payments).",
    ).strip() or default_role

    name_lower = f.name.lower()
    sheet_name: Optional[str] = None

    if name_lower.endswith(".xlsx"):
        xlsx_bytes = f.getvalue()
        try:
            wb_tmp = load_workbook_bytes(xlsx_bytes, data_only=(formula_mode == "Cached values (recommended)"))
            sheets = wb_tmp.sheetnames
        except Exception as e:
            st.error(f"Could not read workbook '{f.name}'. Error: {e}")
            st.stop()

        c1, c2 = st.columns([2, 1])
        with c1:
            sheet_name = st.selectbox(f"Sheet ({f.name})", sheets, key=f"sheet_{i}")
        with c2:
            hdr = st.number_input(f"Header row ({f.name})", 1, 100, int(header_row), 1, key=f"hdr_{i}")
    else:
        hdr = 1  # not used for CSV

    default_key = str(tmpl_defaults.get("merge_keys_by_role", {}).get(role, "AppID"))
    key_col = st.text_input(
        f"Merge key column ({f.name})",
        value=default_key,
        key=f"key_{i}",
        help="Typically AppID. This is used to merge all uploads into one dataset.",
    ).strip() or default_key

    try:
        df = read_uploaded_file_as_df(
            file_obj=f,
            sheet_name=sheet_name,
            header_row=int(hdr),
            formula_mode=formula_mode,
            drop_empty=drop_empty,
        )
    except Exception as e:
        st.error(f"Failed to read '{f.name}'. Error: {e}")
        st.stop()

    df = force_columns_to_text(df, force_text_cols)

    st.caption(f"Rows: {len(df):,} | Columns: {df.shape[1]:,}")
    st.dataframe(df.head(min(preview_rows, 25)), use_container_width=True)

    dfs.append(df)
    file_configs.append({"name": f.name, "role": role, "key_col": key_col, "columns": list(df.columns)})

roles = [cfg["role"] for cfg in file_configs]
if len(set(roles)) != len(roles):
    st.error("Role names must be unique. Please adjust roles so each file has a distinct role.")
    st.stop()

st.divider()
st.subheader("2) Merge")

base_role_default = tmpl_defaults.get("base_role", "")
if base_role_default not in roles:
    base_role_default = roles[0]

base_role = st.selectbox("Base dataset role (others merge into this)", roles, index=roles.index(base_role_default))
base_idx = roles.index(base_role)

join_type_default = tmpl_defaults.get("join_type", "Left (recommended)")
join_type = st.selectbox(
    "Join type",
    ["Left (recommended)", "Inner"],
    index=0 if join_type_default.startswith("Left") else 1,
)
how = "left" if join_type.startswith("Left") else "inner"

st.subheader("Merge quality controls")
exclude_unmatched = st.checkbox(
    "Exclude records not matched across files (recommended)",
    value=True,
    help="Removes records that do not match the base file (prevents orphan rows). Shows a summary of excluded rows.",
)

merged = dfs[base_idx].copy()
base_key = file_configs[base_idx]["key_col"]

if base_key not in merged.columns:
    st.error(f"Base merge key '{base_key}' not found in base file ({file_configs[base_idx]['name']}).")
    st.stop()

# Normalise base key
merged[base_key] = normalise_merge_key(merged[base_key])

blank_base = (merged[base_key] == "").sum()
if blank_base > 0:
    st.warning(f"Base dataset has {blank_base:,} blank '{base_key}' keys. These rows will not match other files.")

exclusion_notes = []  # capture messages for the UI

for j in range(len(dfs)):
    if j == base_idx:
        continue

    df_j = dfs[j].copy()
    key_j = file_configs[j]["key_col"]
    role_j = file_configs[j]["role"]
    file_j = file_configs[j]["name"]

    if key_j not in df_j.columns:
        st.error(f"Merge key '{key_j}' not found in file: {file_j}")
        st.stop()

    # Normalise merge key on the right
    df_j[key_j] = normalise_merge_key(df_j[key_j])

    # Drop blank keys in the right-hand file (prevents keyless rows)
    before = len(df_j)
    df_j = df_j[df_j[key_j] != ""]
    dropped_blank = before - len(df_j)
    if dropped_blank > 0:
        exclusion_notes.append(f"{file_j}: dropped {dropped_blank:,} rows with blank key '{key_j}'")

    # De-duplicate right keys to avoid row multiplication
    if df_j[key_j].duplicated().any():
        dup_n = int(df_j[key_j].duplicated().sum())
        exclusion_notes.append(f"{file_j}: {dup_n:,} duplicate keys in '{key_j}' (kept first occurrence)")
        df_j = df_j.drop_duplicates(subset=[key_j], keep="first")

    # Merge with indicator to detect unmatched
    ind_col = f"__merge__{safe_filename(role_j, 24)}"
    merged = merged.merge(
        df_j,
        left_on=base_key,
        right_on=key_j,
        how=how,
        suffixes=("", f"__{safe_filename(role_j, 24)}"),
        indicator=ind_col,
    )

    # If requested, exclude unmatched rows
    if exclude_unmatched:
        # For LEFT join, "right_only" should not happen; for safety and for INNER/other edge cases, handle anyway.
        right_only = (merged[ind_col] == "right_only").sum()
        left_only = (merged[ind_col] == "left_only").sum()

        if right_only > 0:
            exclusion_notes.append(f"{file_j}: excluded {right_only:,} orphan rows (present only in this file)")
            merged = merged[merged[ind_col] != "right_only"].copy()

        # Note: left_only rows are base rows with no match in this file.
        # We do NOT drop these by default because base is authoritative, but we report them.
        if left_only > 0:
            exclusion_notes.append(f"{file_j}: {left_only:,} base records had no match in this file")

    # Drop indicator column (keeps dataset clean)
    merged = merged.drop(columns=[ind_col], errors="ignore")

# Display summary
if exclusion_notes:
    st.info("Merge notes:\n\n- " + "\n- ".join(exclusion_notes))

st.success(f"Merged rows: {len(merged):,} | Columns: {merged.shape[1]:,}")
st.dataframe(merged.head(preview_rows), use_container_width=True)

st.divider()
st.subheader("3) Build export columns (map + transform + rename)")

all_cols = list(merged.columns)


tmpl_out: List[Dict[str, Any]] = list(tmpl_defaults.get("output_spec", [])) if tmpl_defaults else []
default_num = max(10, len(tmpl_out))

num_out = st.number_input("How many output columns?", min_value=1, max_value=200, value=int(default_num), step=1)

out_rows: List[Dict[str, Any]] = []

for k in range(int(num_out)):
    d_src = "(blank)"
    d_tf = "None"
    d_out = ""
    d_params: Dict[str, Any] = {}

    if k < len(tmpl_out):
        d_src = str(tmpl_out[k].get("source", "(blank)"))
        d_tf = str(tmpl_out[k].get("transform", "None"))
        d_out = str(tmpl_out[k].get("output_name", ""))
        d_params = dict(tmpl_out[k].get("params", {}) or {})

    if d_tf not in TRANSFORM_FUNCS:
        d_tf = "None"

    c1, c2, c3 = st.columns([3, 2, 3])
    with c1:
        src = st.selectbox(
            f"Source column #{k+1}",
            options=["(blank)"] + all_cols,
            index=(["(blank)"] + all_cols).index(d_src) if d_src in (["(blank)"] + all_cols) else 0,
            key=f"src_{k}",
        )

    with c2:
        tf = st.selectbox(
            f"Transform #{k+1}",
            options=list(TRANSFORM_FUNCS.keys()),
            index=list(TRANSFORM_FUNCS.keys()).index(d_tf),
            key=f"tf_{k}",
        )

    with c3:
        default_out_name = d_out if d_out else ("" if src == "(blank)" else str(src))
        out_name = st.text_input(
            f"Output column name #{k+1}",
            value=default_out_name,
            key=f"out_{k}",
        )

    params: Dict[str, Any] = {}
    if tf == "Digits: keep last N":
        params["n"] = st.number_input(
            f"N (digits) #{k+1}",
            min_value=1,
            max_value=50,
            value=int(d_params.get("n", 4)),
            step=1,
            key=f"n_{k}",
        )
    elif tf == "Extract by regex":
        params["pattern"] = st.text_input(
            f"Regex pattern #{k+1}",
            value=str(d_params.get("pattern", r"(\w+)")),
            key=f"rx_{k}",
        )
        params["group"] = st.number_input(
            f"Regex group #{k+1}",
            min_value=0,
            max_value=20,
            value=int(d_params.get("group", 1)),
            step=1,
            key=f"grp_{k}",
        )
        params["ignore_case"] = st.checkbox(
            f"Ignore case #{k+1}",
            value=bool(d_params.get("ignore_case", True)),
            key=f"ic_{k}",
        )
    elif tf == "Split + take part":
        params["delim"] = st.text_input(
            f"Delimiter #{k+1}",
            value=str(d_params.get("delim", ",")),
            key=f"delim_{k}",
        )
        params["index"] = st.number_input(
            f"Index (0-based; -1=last) #{k+1}",
            min_value=-50,
            max_value=50,
            value=int(d_params.get("index", 0)),
            step=1,
            key=f"idx_{k}",
        )
    elif tf == "Prefix if missing":
        params["prefix"] = st.text_input(
            f"Prefix #{k+1}",
            value=str(d_params.get("prefix", "")),
            key=f"pre_{k}",
        )
    elif tf == "Suffix":
        params["suffix"] = st.text_input(
            f"Suffix #{k+1}",
            value=str(d_params.get("suffix", "")),
            key=f"suf_{k}",
        )
    elif tf == "Regex replace":
        params["pattern"] = st.text_input(
            f"Regex pattern #{k+1}",
            value=str(d_params.get("pattern", r"\s+")),
            key=f"rrx_{k}",
        )
        params["repl"] = st.text_input(
            f"Replace with #{k+1}",
            value=str(d_params.get("repl", "")),
            key=f"rrepl_{k}",
        )
        params["ignore_case"] = st.checkbox(
            f"Ignore case (replace) #{k+1}",
            value=bool(d_params.get("ignore_case", True)),
            key=f"ric_{k}",
        )

    out_rows.append({"source": src, "transform": tf, "params": params, "output_name": out_name.strip()})

export_df = pd.DataFrame()

for row in out_rows:
    src = row["source"]
    out_col = row["output_name"]
    tf = row["transform"]
    params = row.get("params", {}) or {}

    if src == "(blank)" or not out_col:
        continue
    if src not in merged.columns:
        continue

    try:
        export_df[out_col] = apply_transform(merged[src], tf, params)
    except Exception:
        export_df[out_col] = apply_transform(merged[src].astype(str), tf, params)

st.subheader("Export preview")
st.caption(f"Rows: {len(export_df):,} | Columns: {export_df.shape[1]:,}")
st.dataframe(export_df.head(preview_rows), use_container_width=True)

st.divider()
st.subheader("4) Download")

timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
base_name = safe_filename(base_role)

out_name = f"{base_name}_{timestamp}.csv"
try:
    csv_bytes = to_csv_bytes(
        df=export_df,
        delimiter=delimiter,
        encoding=encoding_map[encoding],
        quoting=quoting,
        escapechar_enabled=escapechar_enabled,
        date_format=date_format,
    )
    st.download_button(
        label="Download Merged/Transformed CSV",
        data=csv_bytes,
        file_name=out_name,
        mime="text/csv",
    )
except Exception as e:
    st.error(f"Failed to create CSV. Error: {e}")

st.divider()
st.subheader("5) Save mapping template (future-proof)")

merge_keys_by_role = {cfg["role"]: cfg["key_col"] for cfg in file_configs}

template_payload = build_template_payload(
    join_type=join_type,
    base_role=base_role,
    merge_keys_by_role=merge_keys_by_role,
    out_rows=out_rows,
)

tmpl_json = json.dumps(template_payload, indent=2).encode("utf-8")
st.download_button(
    "Download mapping template (.json)",
    data=tmpl_json,
    file_name=f"{safe_filename(base_role)}_mapping_template.json",
    mime="application/json",
)

st.caption(
    "Tip: Load the template next time to pre-fill merge keys and output mapping. "
    "If a transform name is missing in a future version, it will safely fall back to 'None'."
)
